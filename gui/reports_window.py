#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Reports Window for Mobile Shop Management System
نافذة التقارير لنظام إدارة محل الموبايلات
"""

import customtkinter as ctk
from tkinter import messagebox, filedialog
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.dates as mdates
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import sys
import os

# Add project root to path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Import Arabic font support
from utils.arabic_support import create_title_font, create_heading_font, create_button_font, create_body_font

class ReportsWindow(ctk.CTkFrame):
    """Reports and analytics window"""

    def __init__(self, parent, db_manager):
        super().__init__(parent)
        self.db_manager = db_manager

        # Configure matplotlib for Arabic support
        from utils.font_loader import configure_matplotlib_arabic
        configure_matplotlib_arabic()

        self.create_widgets()
        self.load_default_report()

    def create_widgets(self):
        """Create the reports interface"""
        # Title
        self.title_label = ctk.CTkLabel(
            self,
            text="التقارير والإحصائيات",
            font=create_title_font(28)
        )
        self.title_label.pack(pady=(0, 20))

        # Main container
        self.main_container = ctk.CTkFrame(self)
        self.main_container.pack(fill="both", expand=True)

        # Configure grid
        self.main_container.grid_columnconfigure(0, weight=1)
        self.main_container.grid_rowconfigure(1, weight=1)

        # Controls frame
        self.create_controls()

        # Content frame
        self.content_frame = ctk.CTkFrame(self.main_container)
        self.content_frame.grid(row=1, column=0, padx=0, pady=(10, 0), sticky="nsew")

        # Create tabview for different reports
        self.tabview = ctk.CTkTabview(self.content_frame)
        self.tabview.pack(fill="both", expand=True, padx=20, pady=20)

        # Add tabs
        self.sales_tab = self.tabview.add("تقرير المبيعات")
        self.products_tab = self.tabview.add("تقرير المنتجات")
        self.customers_tab = self.tabview.add("تقرير العملاء")
        self.financial_tab = self.tabview.add("التقرير المالي")

        # Create tab contents
        self.create_sales_tab()
        self.create_products_tab()
        self.create_customers_tab()
        self.create_financial_tab()

    def create_controls(self):
        """Create report controls"""
        self.controls_frame = ctk.CTkFrame(self.main_container)
        self.controls_frame.grid(row=0, column=0, padx=0, pady=0, sticky="ew")

        # Date range selection
        self.date_frame = ctk.CTkFrame(self.controls_frame)
        self.date_frame.pack(side="left", padx=20, pady=20)

        self.date_label = ctk.CTkLabel(
            self.date_frame,
            text="نطاق التاريخ:",
            font=create_heading_font(14)
        )
        self.date_label.pack(anchor="w", padx=10, pady=(10, 5))

        # Date inputs frame
        self.date_inputs_frame = ctk.CTkFrame(self.date_frame)
        self.date_inputs_frame.pack(padx=10, pady=(0, 10))

        # From date
        self.from_date_label = ctk.CTkLabel(self.date_inputs_frame, text="من:", font=create_body_font(12))
        self.from_date_label.pack(side="left", padx=(10, 5), pady=10)

        self.from_date_entry = ctk.CTkEntry(
            self.date_inputs_frame,
            placeholder_text="YYYY-MM-DD",
            width=120,
            font=create_body_font(12)
        )
        self.from_date_entry.pack(side="left", padx=5, pady=10)

        # To date
        self.to_date_label = ctk.CTkLabel(self.date_inputs_frame, text="إلى:", font=create_body_font(12))
        self.to_date_label.pack(side="left", padx=(10, 5), pady=10)

        self.to_date_entry = ctk.CTkEntry(
            self.date_inputs_frame,
            placeholder_text="YYYY-MM-DD",
            width=120
        )
        self.to_date_entry.pack(side="left", padx=5, pady=10)

        # Quick date buttons
        self.quick_dates_frame = ctk.CTkFrame(self.date_frame)
        self.quick_dates_frame.pack(padx=10, pady=(0, 10))

        self.today_btn = ctk.CTkButton(
            self.quick_dates_frame,
            text="اليوم",
            command=self.set_today,
            width=80,
            height=30
        )
        self.today_btn.pack(side="left", padx=5, pady=5)

        self.week_btn = ctk.CTkButton(
            self.quick_dates_frame,
            text="هذا الأسبوع",
            command=self.set_this_week,
            width=100,
            height=30
        )
        self.week_btn.pack(side="left", padx=5, pady=5)

        self.month_btn = ctk.CTkButton(
            self.quick_dates_frame,
            text="هذا الشهر",
            command=self.set_this_month,
            width=100,
            height=30
        )
        self.month_btn.pack(side="left", padx=5, pady=5)

        # Action buttons
        self.actions_frame = ctk.CTkFrame(self.controls_frame)
        self.actions_frame.pack(side="right", padx=20, pady=20)

        self.refresh_btn = ctk.CTkButton(
            self.actions_frame,
            text="تحديث التقارير",
            command=self.refresh_reports,
            font=ctk.CTkFont(size=14),
            width=150,
            height=40
        )
        self.refresh_btn.pack(padx=10, pady=10)

        self.export_btn = ctk.CTkButton(
            self.actions_frame,
            text="تصدير إلى Excel",
            command=self.export_to_excel,
            font=ctk.CTkFont(size=14),
            width=150,
            height=40,
            fg_color="green",
            hover_color="darkgreen"
        )
        self.export_btn.pack(padx=10, pady=(0, 10))

    def create_sales_tab(self):
        """Create sales report tab"""
        # Sales summary frame
        self.sales_summary_frame = ctk.CTkFrame(self.sales_tab)
        self.sales_summary_frame.pack(fill="x", padx=20, pady=20)

        self.sales_summary_title = ctk.CTkLabel(
            self.sales_summary_frame,
            text="ملخص المبيعات",
            font=ctk.CTkFont(size=18, weight="bold")
        )
        self.sales_summary_title.pack(pady=(15, 10))

        # Summary stats
        self.sales_stats_frame = ctk.CTkFrame(self.sales_summary_frame)
        self.sales_stats_frame.pack(fill="x", padx=15, pady=(0, 15))

        # Configure grid
        for i in range(4):
            self.sales_stats_frame.grid_columnconfigure(i, weight=1)

        # Create summary cards
        self.total_sales_card = self.create_summary_card(
            self.sales_stats_frame, "إجمالي المبيعات", "0.00 ريال", 0, 0
        )
        self.total_transactions_card = self.create_summary_card(
            self.sales_stats_frame, "عدد المعاملات", "0", 0, 1
        )
        self.avg_transaction_card = self.create_summary_card(
            self.sales_stats_frame, "متوسط المعاملة", "0.00 ريال", 0, 2
        )
        self.total_items_card = self.create_summary_card(
            self.sales_stats_frame, "العناصر المباعة", "0", 0, 3
        )

        # Sales chart
        self.sales_chart_frame = ctk.CTkFrame(self.sales_tab)
        self.sales_chart_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        self.sales_chart_title = ctk.CTkLabel(
            self.sales_chart_frame,
            text="اتجاه المبيعات",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        self.sales_chart_title.pack(pady=(15, 10))

    def create_products_tab(self):
        """Create products report tab"""
        # Top products frame
        self.top_products_frame = ctk.CTkFrame(self.products_tab)
        self.top_products_frame.pack(fill="x", padx=20, pady=20)

        self.top_products_title = ctk.CTkLabel(
            self.top_products_frame,
            text="أفضل المنتجات مبيعاً",
            font=ctk.CTkFont(size=18, weight="bold")
        )
        self.top_products_title.pack(pady=(15, 10))

        # Top products list
        self.top_products_scroll = ctk.CTkScrollableFrame(self.top_products_frame, height=200)
        self.top_products_scroll.pack(fill="x", padx=15, pady=(0, 15))

        # Stock levels frame
        self.stock_frame = ctk.CTkFrame(self.products_tab)
        self.stock_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        self.stock_title = ctk.CTkLabel(
            self.stock_frame,
            text="حالة المخزون",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        self.stock_title.pack(pady=(15, 10))

    def create_customers_tab(self):
        """Create customers report tab"""
        # Customer stats frame
        self.customer_stats_frame = ctk.CTkFrame(self.customers_tab)
        self.customer_stats_frame.pack(fill="x", padx=20, pady=20)

        self.customer_stats_title = ctk.CTkLabel(
            self.customer_stats_frame,
            text="إحصائيات العملاء",
            font=ctk.CTkFont(size=18, weight="bold")
        )
        self.customer_stats_title.pack(pady=(15, 10))

        # Stats grid
        self.customer_stats_grid = ctk.CTkFrame(self.customer_stats_frame)
        self.customer_stats_grid.pack(fill="x", padx=15, pady=(0, 15))

        # Configure grid
        for i in range(3):
            self.customer_stats_grid.grid_columnconfigure(i, weight=1)

        # Customer stats cards
        self.total_customers_card = self.create_summary_card(
            self.customer_stats_grid, "إجمالي العملاء", "0", 0, 0
        )
        self.new_customers_card = self.create_summary_card(
            self.customer_stats_grid, "عملاء جدد", "0", 0, 1
        )
        self.active_customers_card = self.create_summary_card(
            self.customer_stats_grid, "عملاء نشطون", "0", 0, 2
        )

        # Top customers frame
        self.top_customers_frame = ctk.CTkFrame(self.customers_tab)
        self.top_customers_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        self.top_customers_title = ctk.CTkLabel(
            self.top_customers_frame,
            text="أفضل العملاء",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        self.top_customers_title.pack(pady=(15, 10))

        self.top_customers_scroll = ctk.CTkScrollableFrame(self.top_customers_frame)
        self.top_customers_scroll.pack(fill="both", expand=True, padx=15, pady=(0, 15))

    def create_financial_tab(self):
        """Create financial report tab"""
        # Financial summary frame
        self.financial_summary_frame = ctk.CTkFrame(self.financial_tab)
        self.financial_summary_frame.pack(fill="x", padx=20, pady=20)

        self.financial_summary_title = ctk.CTkLabel(
            self.financial_summary_frame,
            text="الملخص المالي",
            font=ctk.CTkFont(size=18, weight="bold")
        )
        self.financial_summary_title.pack(pady=(15, 10))

        # Financial stats
        self.financial_stats_frame = ctk.CTkFrame(self.financial_summary_frame)
        self.financial_stats_frame.pack(fill="x", padx=15, pady=(0, 15))

        # Configure grid
        for i in range(3):
            self.financial_stats_frame.grid_columnconfigure(i, weight=1)

        # Financial cards
        self.revenue_card = self.create_summary_card(
            self.financial_stats_frame, "إجمالي الإيرادات", "0.00 ريال", 0, 0
        )
        self.cost_card = self.create_summary_card(
            self.financial_stats_frame, "إجمالي التكلفة", "0.00 ريال", 0, 1
        )
        self.profit_card = self.create_summary_card(
            self.financial_stats_frame, "صافي الربح", "0.00 ريال", 0, 2
        )

        # Profit chart frame
        self.profit_chart_frame = ctk.CTkFrame(self.financial_tab)
        self.profit_chart_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        self.profit_chart_title = ctk.CTkLabel(
            self.profit_chart_frame,
            text="اتجاه الأرباح",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        self.profit_chart_title.pack(pady=(15, 10))

    def create_summary_card(self, parent, title, value, row, col):
        """Create a summary card widget"""
        card_frame = ctk.CTkFrame(parent)
        card_frame.grid(row=row, column=col, padx=10, pady=10, sticky="ew")

        value_label = ctk.CTkLabel(
            card_frame,
            text=value,
            font=ctk.CTkFont(size=20, weight="bold")
        )
        value_label.pack(pady=(15, 5))

        title_label = ctk.CTkLabel(
            card_frame,
            text=title,
            font=ctk.CTkFont(size=12)
        )
        title_label.pack(pady=(5, 15))

        return {'value': value_label, 'title': title_label}

    def set_today(self):
        """Set date range to today"""
        today = datetime.now().strftime('%Y-%m-%d')
        self.from_date_entry.delete(0, "end")
        self.from_date_entry.insert(0, today)
        self.to_date_entry.delete(0, "end")
        self.to_date_entry.insert(0, today)

    def set_this_week(self):
        """Set date range to this week"""
        today = datetime.now()
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=6)

        self.from_date_entry.delete(0, "end")
        self.from_date_entry.insert(0, start_of_week.strftime('%Y-%m-%d'))
        self.to_date_entry.delete(0, "end")
        self.to_date_entry.insert(0, end_of_week.strftime('%Y-%m-%d'))

    def set_this_month(self):
        """Set date range to this month"""
        today = datetime.now()
        start_of_month = today.replace(day=1)

        # Get last day of month
        if today.month == 12:
            end_of_month = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            end_of_month = today.replace(month=today.month + 1, day=1) - timedelta(days=1)

        self.from_date_entry.delete(0, "end")
        self.from_date_entry.insert(0, start_of_month.strftime('%Y-%m-%d'))
        self.to_date_entry.delete(0, "end")
        self.to_date_entry.insert(0, end_of_month.strftime('%Y-%m-%d'))

    def load_default_report(self):
        """Load default report (this month)"""
        self.set_this_month()
        self.refresh_reports()

    def get_date_range(self):
        """Get the selected date range"""
        start_date = self.from_date_entry.get().strip()
        end_date = self.to_date_entry.get().strip()

        # Validate dates
        try:
            if start_date:
                datetime.strptime(start_date, '%Y-%m-%d')
            if end_date:
                datetime.strptime(end_date, '%Y-%m-%d')
        except ValueError:
            messagebox.showerror("خطأ", "تنسيق التاريخ غير صحيح. استخدم YYYY-MM-DD")
            return None, None

        return start_date or None, end_date or None

    def refresh_reports(self):
        """Refresh all reports with current date range"""
        start_date, end_date = self.get_date_range()
        if start_date is None and end_date is None:
            return

        try:
            self.update_sales_report(start_date, end_date)
            self.update_products_report(start_date, end_date)
            self.update_customers_report(start_date, end_date)
            self.update_financial_report(start_date, end_date)
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ في تحديث التقارير: {e}")

    def update_sales_report(self, start_date, end_date):
        """Update sales report"""
        # Get sales data
        sales_data = self.db_manager.get_sales_report(start_date, end_date)

        # Calculate summary statistics with None value handling
        total_sales = 0
        for sale in sales_data:
            amount = sale.get('total_amount')
            if amount is not None:
                try:
                    total_sales += float(amount)
                except (ValueError, TypeError):
                    continue
        
        total_transactions = len(sales_data)
        avg_transaction = total_sales / total_transactions if total_transactions > 0 else 0

        # Get total items sold
        items_query = '''
            SELECT COALESCE(SUM(si.quantity), 0) as total_items
            FROM sale_items si
            JOIN sales s ON si.sale_id = s.id
        '''
        params = []
        if start_date and end_date:
            items_query += " WHERE DATE(s.sale_date) BETWEEN ? AND ?"
            params = [start_date, end_date]
        elif start_date:
            items_query += " WHERE DATE(s.sale_date) >= ?"
            params = [start_date]
        elif end_date:
            items_query += " WHERE DATE(s.sale_date) <= ?"
            params = [end_date]

        items_result = self.db_manager.execute_query(items_query, tuple(params))
        total_items = items_result[0]['total_items'] if items_result else 0

        # Update summary cards
        self.total_sales_card['value'].configure(text=f"{total_sales:.2f} ريال")
        self.total_transactions_card['value'].configure(text=str(total_transactions))
        self.avg_transaction_card['value'].configure(text=f"{avg_transaction:.2f} ريال")
        self.total_items_card['value'].configure(text=str(total_items))

        # Update sales chart
        self.update_sales_chart(start_date, end_date)

    def update_sales_chart(self, start_date, end_date):
        """Update sales trend chart"""
        # Clear previous chart
        for widget in self.sales_chart_frame.winfo_children():
            if isinstance(widget, FigureCanvasTkAgg):
                widget.get_tk_widget().destroy()

        # Get daily sales data
        query = '''
            SELECT DATE(sale_date) as sale_day, 
                   COALESCE(SUM(total_amount), 0) as daily_total
            FROM sales
        '''
        params = []
        if start_date and end_date:
            query += " WHERE DATE(sale_date) BETWEEN ? AND ?"
            params = [start_date, end_date]
        elif start_date:
            query += " WHERE DATE(sale_date) >= ?"
            params = [start_date]
        elif end_date:
            query += " WHERE DATE(sale_date) <= ?"
            params = [end_date]

        query += " GROUP BY DATE(sale_date) ORDER BY sale_day"

        daily_sales = self.db_manager.execute_query(query, tuple(params))

        # Create chart
        fig, ax = plt.subplots(figsize=(10, 4))
        fig.patch.set_facecolor('none')
        ax.set_facecolor('none')

        if daily_sales:
            dates = []
            amounts = []
            for row in daily_sales:
                try:
                    dates.append(datetime.strptime(row['sale_day'], '%Y-%m-%d'))
                    amount = row.get('daily_total', 0)
                    amounts.append(float(amount) if amount is not None else 0)
                except (ValueError, TypeError) as e:
                    print(f"خطأ في معالجة البيانات: {e}")
                    continue

            ax.plot(dates, amounts, marker='o', linewidth=2, markersize=4)
            ax.fill_between(dates, amounts, alpha=0.3)

            # Format x-axis
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%d'))
            if len(dates) > 10:
                ax.xaxis.set_major_locator(mdates.DayLocator(interval=max(1, len(dates)//10)))

            plt.xticks(rotation=45)
        else:
            ax.text(0.5, 0.5, 'لا توجد بيانات للفترة المحددة', ha='center', va='center', 
                   transform=ax.transAxes, fontsize=14)

        ax.set_xlabel('التاريخ')
        ax.set_ylabel('المبيعات (ريال)')
        ax.set_title('اتجاه المبيعات اليومية')

        plt.tight_layout()

        # Add chart to frame
        canvas = FigureCanvasTkAgg(fig, self.sales_chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True, padx=15, pady=(0, 15))

    def update_products_report(self, start_date, end_date):
        """Update products report"""
        # Clear previous data
        for widget in self.top_products_scroll.winfo_children():
            widget.destroy()

        # Get top selling products
        query = '''
            SELECT p.name, p.brand, SUM(si.quantity) as total_sold, 
                   SUM(si.total_price) as total_revenue
            FROM sale_items si
            JOIN products p ON si.product_id = p.id
            JOIN sales s ON si.sale_id = s.id
        '''
        params = []
        if start_date and end_date:
            query += " WHERE DATE(s.sale_date) BETWEEN ? AND ?"
            params = [start_date, end_date]
        elif start_date:
            query += " WHERE DATE(s.sale_date) >= ?"
            params = [start_date]
        elif end_date:
            query += " WHERE DATE(s.sale_date) <= ?"
            params = [end_date]

        query += '''
            GROUP BY p.id, p.name, p.brand
            ORDER BY total_sold DESC
            LIMIT 10
        '''

        top_products = self.db_manager.execute_query(query, tuple(params))

        if not top_products:
            no_data_label = ctk.CTkLabel(
                self.top_products_scroll,
                text="لا توجد بيانات للفترة المحددة",
                font=ctk.CTkFont(size=14)
            )
            no_data_label.pack(pady=20)
        else:
            for i, product in enumerate(top_products, 1):
                product_frame = ctk.CTkFrame(self.top_products_scroll)
                product_frame.pack(fill="x", padx=5, pady=2)

                info_label = ctk.CTkLabel(
                    product_frame,
                    text=f"{i}. {product['name']} ({product['brand']}) - "
                         f"الكمية: {product['total_sold']} | الإيرادات: {product['total_revenue']:.2f} ريال",
                    font=ctk.CTkFont(size=12),
                    anchor="w"
                )
                info_label.pack(fill="x", padx=10, pady=8)

        # Update stock levels chart
        self.update_stock_chart()

    def update_stock_chart(self):
        """Update stock levels chart"""
        # Clear previous chart
        for widget in self.stock_frame.winfo_children():
            if isinstance(widget, FigureCanvasTkAgg):
                widget.get_tk_widget().destroy()

        # Get stock data
        stock_data = self.db_manager.execute_query('''
            SELECT 
                SUM(CASE WHEN quantity > min_quantity THEN 1 ELSE 0 END) as good_stock,
                SUM(CASE WHEN quantity <= min_quantity AND quantity > 0 THEN 1 ELSE 0 END) as low_stock,
                SUM(CASE WHEN quantity = 0 THEN 1 ELSE 0 END) as out_of_stock
            FROM products
        ''')

        if stock_data:
            data = stock_data[0]
            labels = ['مخزون جيد', 'مخزون منخفض', 'نفد المخزون']
            values = [data['good_stock'], data['low_stock'], data['out_of_stock']]
            colors = ['green', 'orange', 'red']

            # Create pie chart
            fig, ax = plt.subplots(figsize=(6, 4))
            fig.patch.set_facecolor('none')

            if sum(values) > 0:
                wedges, texts, autotexts = ax.pie(values, labels=labels, colors=colors, 
                                                autopct='%1.1f%%', startangle=90)
                ax.set_title('توزيع حالة المخزون')
            else:
                ax.text(0.5, 0.5, 'لا توجد منتجات', ha='center', va='center')

            plt.tight_layout()

            # Add chart to frame
            canvas = FigureCanvasTkAgg(fig, self.stock_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True, padx=15, pady=(0, 15))

    def update_customers_report(self, start_date, end_date):
        """Update customers report"""
        # Get customer statistics
        total_customers = self.db_manager.execute_query("SELECT COUNT(*) as count FROM customers")[0]['count']

        # Get new customers in date range
        new_customers_query = "SELECT COUNT(*) as count FROM customers"
        new_params = []
        if start_date and end_date:
            new_customers_query += " WHERE DATE(created_at) BETWEEN ? AND ?"
            new_params = [start_date, end_date]
        elif start_date:
            new_customers_query += " WHERE DATE(created_at) >= ?"
            new_params = [start_date]
        elif end_date:
            new_customers_query += " WHERE DATE(created_at) <= ?"
            new_params = [end_date]

        new_customers = self.db_manager.execute_query(new_customers_query, tuple(new_params))[0]['count']

        # Get active customers (customers with purchases in date range)
        active_customers_query = '''
            SELECT COUNT(DISTINCT customer_id) as count FROM sales 
            WHERE customer_id IS NOT NULL
        '''
        active_params = []
        if start_date and end_date:
            active_customers_query += " AND DATE(sale_date) BETWEEN ? AND ?"
            active_params = [start_date, end_date]
        elif start_date:
            active_customers_query += " AND DATE(sale_date) >= ?"
            active_params = [start_date]
        elif end_date:
            active_customers_query += " AND DATE(sale_date) <= ?"
            active_params = [end_date]

        active_customers = self.db_manager.execute_query(active_customers_query, tuple(active_params))[0]['count']

        # Update customer stats cards
        self.total_customers_card['value'].configure(text=str(total_customers))
        self.new_customers_card['value'].configure(text=str(new_customers))
        self.active_customers_card['value'].configure(text=str(active_customers))

        # Get top customers
        self.update_top_customers(start_date, end_date)

    def update_top_customers(self, start_date, end_date):
        """Update top customers list"""
        # Clear previous data
        for widget in self.top_customers_scroll.winfo_children():
            widget.destroy()

        # Get top customers by purchase amount
        query = '''
            SELECT c.name, c.phone, COALESCE(SUM(s.total_amount), 0) as total_spent,
                   COUNT(s.id) as purchase_count
            FROM customers c
            LEFT JOIN sales s ON c.id = s.customer_id
        '''
        params = []
        if start_date and end_date:
            query += " AND DATE(s.sale_date) BETWEEN ? AND ?"
            params = [start_date, end_date]
        elif start_date:
            query += " AND DATE(s.sale_date) >= ?"
            params = [start_date]
        elif end_date:
            query += " AND DATE(s.sale_date) <= ?"
            params = [end_date]

        query += '''
            GROUP BY c.id, c.name, c.phone
            HAVING total_spent > 0
            ORDER BY total_spent DESC
            LIMIT 10
        '''

        top_customers = self.db_manager.execute_query(query, tuple(params))

        if not top_customers:
            no_data_label = ctk.CTkLabel(
                self.top_customers_scroll,
                text="لا توجد بيانات للفترة المحددة",
                font=ctk.CTkFont(size=14)
            )
            no_data_label.pack(pady=20)
        else:
            for i, customer in enumerate(top_customers, 1):
                customer_frame = ctk.CTkFrame(self.top_customers_scroll)
                customer_frame.pack(fill="x", padx=5, pady=2)

                info_label = ctk.CTkLabel(
                    customer_frame,
                    text=f"{i}. {customer['name']} ({customer['phone']}) - "
                         f"المبلغ: {customer['total_spent']:.2f} ريال | المشتريات: {customer['purchase_count']}",
                    font=ctk.CTkFont(size=12),
                    anchor="w"
                )
                info_label.pack(fill="x", padx=10, pady=8)

    def update_financial_report(self, start_date, end_date):
        """Update financial report"""
        # Get financial data
        financial_query = '''
            SELECT 
                COALESCE(SUM(s.total_amount), 0) as total_revenue,
                COALESCE(SUM(si.quantity * p.purchase_price), 0) as total_cost
            FROM sales s
            JOIN sale_items si ON s.id = si.sale_id
            JOIN products p ON si.product_id = p.id
        '''
        params = []
        if start_date and end_date:
            financial_query += " WHERE DATE(s.sale_date) BETWEEN ? AND ?"
            params = [start_date, end_date]
        elif start_date:
            financial_query += " WHERE DATE(s.sale_date) >= ?"
            params = [start_date]
        elif end_date:
            financial_query += " WHERE DATE(s.sale_date) <= ?"
            params = [end_date]

        financial_data = self.db_manager.execute_query(financial_query, tuple(params))

        if financial_data:
            revenue = float(financial_data[0]['total_revenue'] or 0)
            cost = float(financial_data[0]['total_cost'] or 0)
            profit = revenue - cost

            # Update financial cards
            self.revenue_card['value'].configure(text=f"{revenue:.2f} ريال")
            self.cost_card['value'].configure(text=f"{cost:.2f} ريال")
            self.profit_card['value'].configure(text=f"{profit:.2f} ريال")

            # Set profit card color based on value
            if profit > 0:
                self.profit_card['value'].configure(text_color="green")
            elif profit < 0:
                self.profit_card['value'].configure(text_color="red")
            else:
                self.profit_card['value'].configure(text_color="gray")

        # Update profit chart
        self.update_profit_chart(start_date, end_date)

    def update_profit_chart(self, start_date, end_date):
        """Update profit trend chart"""
        # Clear previous chart
        for widget in self.profit_chart_frame.winfo_children():
            if isinstance(widget, FigureCanvasTkAgg):
                widget.get_tk_widget().destroy()

        # Get daily profit data
        profit_query = '''
            SELECT DATE(s.sale_date) as sale_day,
                   COALESCE(SUM(s.total_amount), 0) as daily_revenue,
                   COALESCE(SUM(si.quantity * p.purchase_price), 0) as daily_cost
            FROM sales s
            JOIN sale_items si ON s.id = si.sale_id
            JOIN products p ON si.product_id = p.id
        '''
        params = []
        if start_date and end_date:
            profit_query += " WHERE DATE(s.sale_date) BETWEEN ? AND ?"
            params = [start_date, end_date]
        elif start_date:
            profit_query += " WHERE DATE(s.sale_date) >= ?"
            params = [start_date]
        elif end_date:
            profit_query += " WHERE DATE(s.sale_date) <= ?"
            params = [end_date]

        profit_query += " GROUP BY DATE(s.sale_date) ORDER BY sale_day"

        profit_data = self.db_manager.execute_query(profit_query, tuple(params))

        # Create chart
        fig, ax = plt.subplots(figsize=(10, 4))
        fig.patch.set_facecolor('none')
        ax.set_facecolor('none')

        if profit_data:
            dates = [datetime.strptime(row['sale_day'], '%Y-%m-%d') for row in profit_data]
            profits = [float(row['daily_revenue'] or 0) - float(row['daily_cost'] or 0) for row in profit_data]

            # Color bars based on profit/loss
            colors = ['green' if p >= 0 else 'red' for p in profits]
            ax.bar(dates, profits, color=colors, alpha=0.7)

            # Format x-axis
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%d'))
            if len(dates) > 10:
                ax.xaxis.set_major_locator(mdates.DayLocator(interval=max(1, len(dates)//10)))

            plt.xticks(rotation=45)
        else:
            ax.text(0.5, 0.5, 'لا توجد بيانات للفترة المحددة', ha='center', va='center', 
                   transform=ax.transAxes, fontsize=14)

        ax.set_xlabel('التاريخ')
        ax.set_ylabel('الربح (ريال)')
        ax.set_title('اتجاه الأرباح اليومية')
        ax.axhline(y=0, color='black', linestyle='-', alpha=0.3)

        plt.tight_layout()

        # Add chart to frame
        canvas = FigureCanvasTkAgg(fig, self.profit_chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True, padx=15, pady=(0, 15))

    def export_to_excel(self):
        """Export current reports to Excel"""
        try:
            # Get date range
            start_date, end_date = self.get_date_range()

            # Choose save location
            filename = filedialog.asksaveasfilename(
                title="حفظ التقرير",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialname=f"mobile_shop_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )

            if not filename:
                return

            # Create Excel workbook
            wb = openpyxl.Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Create sheets
            self.create_sales_excel_sheet(wb, start_date, end_date)
            self.create_products_excel_sheet(wb, start_date, end_date)
            self.create_customers_excel_sheet(wb, start_date, end_date)
            self.create_financial_excel_sheet(wb, start_date, end_date)

            # Save workbook
            wb.save(filename)
            messagebox.showinfo("نجح", f"تم تصدير التقرير بنجاح إلى:\n{filename}")

        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ في تصدير التقرير: {e}")

    def create_sales_excel_sheet(self, wb, start_date, end_date):
        """Create sales sheet in Excel"""
        ws = wb.create_sheet("تقرير المبيعات")

        # Headers style
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Title
        ws['A1'] = "تقرير المبيعات"
        ws['A1'].font = Font(size=16, bold=True)

        # Date range
        date_range = f"من {start_date or 'البداية'} إلى {end_date or 'النهاية'}"
        ws['A2'] = date_range

        # Get sales data
        sales_data = self.db_manager.get_sales_report(start_date, end_date)

        # Headers
        headers = ['رقم الفاتورة', 'العميل', 'المبلغ الإجمالي', 'الخصم', 'الضريبة', 'طريقة الدفع', 'التاريخ']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Data
        for row, sale in enumerate(sales_data, 5):
            ws.cell(row=row, column=1, value=sale['id'])
            ws.cell(row=row, column=2, value=sale.get('customer_name', 'عميل عادي'))
            ws.cell(row=row, column=3, value=sale['total_amount'])
            ws.cell(row=row, column=4, value=sale['discount'])
            ws.cell(row=row, column=5, value=sale['tax'])
            ws.cell(row=row, column=6, value=sale['payment_method'])
            ws.cell(row=row, column=7, value=sale['sale_date'])

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def create_products_excel_sheet(self, wb, start_date, end_date):
        """Create products sheet in Excel"""
        ws = wb.create_sheet("تقرير المنتجات")

        # Similar structure to sales sheet
        # Implementation would follow similar pattern
        ws['A1'] = "تقرير المنتجات"
        ws['A1'].font = Font(size=16, bold=True)

        # Add product statistics and top selling products
        # This is a simplified version
        ws['A3'] = "أفضل المنتجات مبيعاً"
        ws['A3'].font = Font(bold=True)

    def create_customers_excel_sheet(self, wb, start_date, end_date):
        """Create customers sheet in Excel"""
        ws = wb.create_sheet("تقرير العملاء")

        ws['A1'] = "تقرير العملاء"
        ws['A1'].font = Font(size=16, bold=True)

        # Add customer statistics
        ws['A3'] = "إحصائيات العملاء"
        ws['A3'].font = Font(bold=True)

    def create_financial_excel_sheet(self, wb, start_date, end_date):
        """Create financial sheet in Excel"""
        ws = wb.create_sheet("التقرير المالي")

        ws['A1'] = "التقرير المالي"
        ws['A1'].font = Font(size=16, bold=True)

        # Add financial summary
        ws['A3'] = "الملخص المالي"
        ws['A3'].font = Font(bold=True)