
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Settings View
عرض الإعدادات
"""

import customtkinter as ctk
import tkinter as tk
from tkinter import messagebox, filedialog
from pathlib import Path
import json

from src.utils.logger import get_logger

logger = get_logger(__name__)

class SettingsView(ctk.CTkFrame):
    """Settings view for application configuration"""

    def __init__(self, parent, db_manager, settings_manager, theme_manager):
        super().__init__(parent)
        
        self.db_manager = db_manager
        self.settings_manager = settings_manager
        self.theme_manager = theme_manager
        
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        
        self._setup_ui()
        self._load_settings()
        
        logger.info("Settings view initialized")

    def _setup_ui(self):
        """Setup the settings interface"""
        # Header
        header = ctk.CTkFrame(self)
        header.grid(row=0, column=0, sticky="ew", padx=20, pady=(20, 10))
        header.grid_columnconfigure(0, weight=1)

        title_label = ctk.CTkLabel(
            header,
            text="⚙️ إعدادات النظام",
            font=self.theme_manager.get_header_font_config(24, "bold")
        )
        title_label.grid(row=0, column=0, pady=20)

        # Main content with scrollable frame
        self.scrollable_frame = ctk.CTkScrollableFrame(self)
        self.scrollable_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=(0, 20))
        self.scrollable_frame.grid_columnconfigure(0, weight=1)

        # Create settings sections
        self._create_shop_info_section()
        self._create_display_settings_section()
        self._create_business_settings_section()
        self._create_system_settings_section()
        self._create_action_buttons()

    def _create_shop_info_section(self):
        """Create shop information section"""
        # Shop Info Section
        shop_frame = ctk.CTkFrame(self.scrollable_frame)
        shop_frame.grid(row=0, column=0, sticky="ew", pady=(0, 15), padx=10)
        shop_frame.grid_columnconfigure(1, weight=1)

        # Section title
        ctk.CTkLabel(
            shop_frame,
            text="🏪 معلومات المحل",
            font=self.theme_manager.get_header_font_config(18, "bold")
        ).grid(row=0, column=0, columnspan=2, pady=(15, 10), sticky="w", padx=15)

        # Shop name
        ctk.CTkLabel(shop_frame, text="اسم المحل:", font=self.theme_manager.get_font_config(12)).grid(
            row=1, column=0, sticky="w", padx=15, pady=5
        )
        self.shop_name_entry = ctk.CTkEntry(shop_frame, font=self.theme_manager.get_font_config(12))
        self.shop_name_entry.grid(row=1, column=1, sticky="ew", padx=15, pady=5)

        # Owner name
        ctk.CTkLabel(shop_frame, text="اسم المالك:", font=self.theme_manager.get_font_config(12)).grid(
            row=2, column=0, sticky="w", padx=15, pady=5
        )
        self.owner_name_entry = ctk.CTkEntry(shop_frame, font=self.theme_manager.get_font_config(12))
        self.owner_name_entry.grid(row=2, column=1, sticky="ew", padx=15, pady=5)

        # Phone
        ctk.CTkLabel(shop_frame, text="رقم الهاتف:", font=self.theme_manager.get_font_config(12)).grid(
            row=3, column=0, sticky="w", padx=15, pady=5
        )
        self.phone_entry = ctk.CTkEntry(shop_frame, font=self.theme_manager.get_font_config(12))
        self.phone_entry.grid(row=3, column=1, sticky="ew", padx=15, pady=5)

        # Email
        ctk.CTkLabel(shop_frame, text="البريد الإلكتروني:", font=self.theme_manager.get_font_config(12)).grid(
            row=4, column=0, sticky="w", padx=15, pady=5
        )
        self.email_entry = ctk.CTkEntry(shop_frame, font=self.theme_manager.get_font_config(12))
        self.email_entry.grid(row=4, column=1, sticky="ew", padx=15, pady=5)

        # Address
        ctk.CTkLabel(shop_frame, text="العنوان:", font=self.theme_manager.get_font_config(12)).grid(
            row=5, column=0, sticky="w", padx=15, pady=5
        )
        self.address_entry = ctk.CTkEntry(shop_frame, font=self.theme_manager.get_font_config(12))
        self.address_entry.grid(row=5, column=1, sticky="ew", padx=15, pady=5)

        # Tax number
        ctk.CTkLabel(shop_frame, text="الرقم الضريبي:", font=self.theme_manager.get_font_config(12)).grid(
            row=6, column=0, sticky="w", padx=15, pady=5
        )
        self.tax_number_entry = ctk.CTkEntry(shop_frame, font=self.theme_manager.get_font_config(12))
        self.tax_number_entry.grid(row=6, column=1, sticky="ew", padx=15, pady=(5, 15))

    def _create_display_settings_section(self):
        """Create display settings section"""
        display_frame = ctk.CTkFrame(self.scrollable_frame)
        display_frame.grid(row=1, column=0, sticky="ew", pady=(0, 15), padx=10)
        display_frame.grid_columnconfigure(1, weight=1)

        # Section title
        ctk.CTkLabel(
            display_frame,
            text="🎨 إعدادات العرض",
            font=self.theme_manager.get_header_font_config(18, "bold")
        ).grid(row=0, column=0, columnspan=2, pady=(15, 10), sticky="w", padx=15)

        # Theme selection
        ctk.CTkLabel(display_frame, text="السمة:", font=self.theme_manager.get_font_config(12)).grid(
            row=1, column=0, sticky="w", padx=15, pady=5
        )
        self.theme_var = ctk.StringVar()
        self.theme_combo = ctk.CTkComboBox(
            display_frame,
            values=["dark", "light"],
            variable=self.theme_var,
            font=self.theme_manager.get_font_config(12),
            command=self._on_theme_change
        )
        self.theme_combo.grid(row=1, column=1, sticky="ew", padx=15, pady=5)

        # Language selection
        ctk.CTkLabel(display_frame, text="اللغة:", font=self.theme_manager.get_font_config(12)).grid(
            row=2, column=0, sticky="w", padx=15, pady=5
        )
        self.language_var = ctk.StringVar()
        self.language_combo = ctk.CTkComboBox(
            display_frame,
            values=["ar", "en"],
            variable=self.language_var,
            font=self.theme_manager.get_font_config(12)
        )
        self.language_combo.grid(row=2, column=1, sticky="ew", padx=15, pady=5)

        # Font size
        ctk.CTkLabel(display_frame, text="حجم الخط:", font=self.theme_manager.get_font_config(12)).grid(
            row=3, column=0, sticky="w", padx=15, pady=5
        )
        self.font_size_var = ctk.IntVar()
        self.font_size_slider = ctk.CTkSlider(
            display_frame,
            from_=10,
            to=20,
            number_of_steps=10,
            variable=self.font_size_var
        )
        self.font_size_slider.grid(row=3, column=1, sticky="ew", padx=15, pady=5)
        
        self.font_size_label = ctk.CTkLabel(display_frame, text="12", font=self.theme_manager.get_font_config(10))
        self.font_size_label.grid(row=4, column=1, sticky="w", padx=15, pady=(0, 5))

        # Items per page
        ctk.CTkLabel(display_frame, text="عدد العناصر في الصفحة:", font=self.theme_manager.get_font_config(12)).grid(
            row=5, column=0, sticky="w", padx=15, pady=5
        )
        self.items_per_page_var = ctk.StringVar()
        self.items_per_page_combo = ctk.CTkComboBox(
            display_frame,
            values=["25", "50", "100", "200"],
            variable=self.items_per_page_var,
            font=self.theme_manager.get_font_config(12)
        )
        self.items_per_page_combo.grid(row=5, column=1, sticky="ew", padx=15, pady=5)

        # Show grid
        self.show_grid_var = ctk.BooleanVar()
        self.show_grid_check = ctk.CTkCheckBox(
            display_frame,
            text="إظهار الشبكة في الجداول",
            variable=self.show_grid_var,
            font=self.theme_manager.get_font_config(12)
        )
        self.show_grid_check.grid(row=6, column=0, columnspan=2, sticky="w", padx=15, pady=(5, 15))

    def _create_business_settings_section(self):
        """Create business settings section"""
        business_frame = ctk.CTkFrame(self.scrollable_frame)
        business_frame.grid(row=2, column=0, sticky="ew", pady=(0, 15), padx=10)
        business_frame.grid_columnconfigure(1, weight=1)

        # Section title
        ctk.CTkLabel(
            business_frame,
            text="💼 إعدادات الأعمال",
            font=self.theme_manager.get_header_font_config(18, "bold")
        ).grid(row=0, column=0, columnspan=2, pady=(15, 10), sticky="w", padx=15)

        # Currency
        ctk.CTkLabel(business_frame, text="العملة:", font=self.theme_manager.get_font_config(12)).grid(
            row=1, column=0, sticky="w", padx=15, pady=5
        )
        self.currency_entry = ctk.CTkEntry(business_frame, font=self.theme_manager.get_font_config(12))
        self.currency_entry.grid(row=1, column=1, sticky="ew", padx=15, pady=5)

        # Currency symbol
        ctk.CTkLabel(business_frame, text="رمز العملة:", font=self.theme_manager.get_font_config(12)).grid(
            row=2, column=0, sticky="w", padx=15, pady=5
        )
        self.currency_symbol_entry = ctk.CTkEntry(business_frame, font=self.theme_manager.get_font_config(12))
        self.currency_symbol_entry.grid(row=2, column=1, sticky="ew", padx=15, pady=5)

        # Tax rate
        ctk.CTkLabel(business_frame, text="معدل الضريبة (%):", font=self.theme_manager.get_font_config(12)).grid(
            row=3, column=0, sticky="w", padx=15, pady=5
        )
        self.tax_rate_entry = ctk.CTkEntry(business_frame, font=self.theme_manager.get_font_config(12))
        self.tax_rate_entry.grid(row=3, column=1, sticky="ew", padx=15, pady=5)

        # Default discount
        ctk.CTkLabel(business_frame, text="الخصم الافتراضي (%):", font=self.theme_manager.get_font_config(12)).grid(
            row=4, column=0, sticky="w", padx=15, pady=5
        )
        self.default_discount_entry = ctk.CTkEntry(business_frame, font=self.theme_manager.get_font_config(12))
        self.default_discount_entry.grid(row=4, column=1, sticky="ew", padx=15, pady=5)

        # Backup interval
        ctk.CTkLabel(business_frame, text="فترة النسخ الاحتياطي (أيام):", font=self.theme_manager.get_font_config(12)).grid(
            row=5, column=0, sticky="w", padx=15, pady=5
        )
        self.backup_interval_entry = ctk.CTkEntry(business_frame, font=self.theme_manager.get_font_config(12))
        self.backup_interval_entry.grid(row=5, column=1, sticky="ew", padx=15, pady=5)

        # Low stock alert
        self.low_stock_alert_var = ctk.BooleanVar()
        self.low_stock_alert_check = ctk.CTkCheckBox(
            business_frame,
            text="تنبيه المخزون المنخفض",
            variable=self.low_stock_alert_var,
            font=self.theme_manager.get_font_config(12)
        )
        self.low_stock_alert_check.grid(row=6, column=0, columnspan=2, sticky="w", padx=15, pady=(5, 15))

    def _create_system_settings_section(self):
        """Create system settings section"""
        system_frame = ctk.CTkFrame(self.scrollable_frame)
        system_frame.grid(row=3, column=0, sticky="ew", pady=(0, 15), padx=10)
        system_frame.grid_columnconfigure(1, weight=1)

        # Section title
        ctk.CTkLabel(
            system_frame,
            text="🔧 إعدادات النظام",
            font=self.theme_manager.get_header_font_config(18, "bold")
        ).grid(row=0, column=0, columnspan=2, pady=(15, 10), sticky="w", padx=15)

        # Database backup
        backup_button = ctk.CTkButton(
            system_frame,
            text="📁 إنشاء نسخة احتياطية من قاعدة البيانات",
            font=self.theme_manager.get_font_config(12),
            command=self._create_backup
        )
        backup_button.grid(row=1, column=0, columnspan=2, sticky="ew", padx=15, pady=5)

        # Database restore
        restore_button = ctk.CTkButton(
            system_frame,
            text="📥 استعادة نسخة احتياطية",
            font=self.theme_manager.get_font_config(12),
            command=self._restore_backup
        )
        restore_button.grid(row=2, column=0, columnspan=2, sticky="ew", padx=15, pady=5)

        # Export data
        export_button = ctk.CTkButton(
            system_frame,
            text="📊 تصدير البيانات إلى Excel",
            font=self.theme_manager.get_font_config(12),
            command=self._export_data
        )
        export_button.grid(row=3, column=0, columnspan=2, sticky="ew", padx=15, pady=5)

        # Reset settings
        reset_button = ctk.CTkButton(
            system_frame,
            text="🔄 إعادة تعيين الإعدادات",
            font=self.theme_manager.get_font_config(12),
            fg_color="red",
            hover_color="darkred",
            command=self._reset_settings
        )
        reset_button.grid(row=4, column=0, columnspan=2, sticky="ew", padx=15, pady=(5, 15))

    def _create_action_buttons(self):
        """Create action buttons"""
        buttons_frame = ctk.CTkFrame(self.scrollable_frame)
        buttons_frame.grid(row=4, column=0, sticky="ew", pady=(0, 20), padx=10)
        buttons_frame.grid_columnconfigure((0, 1), weight=1)

        # Save button
        save_button = ctk.CTkButton(
            buttons_frame,
            text="💾 حفظ الإعدادات",
            font=self.theme_manager.get_font_config(14, "bold"),
            height=40,
            command=self._save_settings
        )
        save_button.grid(row=0, column=0, sticky="ew", padx=(15, 10), pady=15)

        # Cancel button
        cancel_button = ctk.CTkButton(
            buttons_frame,
            text="❌ إلغاء",
            font=self.theme_manager.get_font_config(14, "bold"),
            height=40,
            fg_color="gray",
            hover_color="darkgray",
            command=self._load_settings
        )
        cancel_button.grid(row=0, column=1, sticky="ew", padx=(10, 15), pady=15)

        # Bind slider change
        self.font_size_var.trace("w", self._on_font_size_change)

    def _load_settings(self):
        """Load current settings into the form"""
        try:
            # Shop info
            self.shop_name_entry.delete(0, "end")
            self.shop_name_entry.insert(0, self.settings_manager.shop_info.name)
            
            self.owner_name_entry.delete(0, "end")
            self.owner_name_entry.insert(0, self.settings_manager.shop_info.owner)
            
            self.phone_entry.delete(0, "end")
            self.phone_entry.insert(0, self.settings_manager.shop_info.phone)
            
            self.email_entry.delete(0, "end")
            self.email_entry.insert(0, self.settings_manager.shop_info.email)
            
            self.address_entry.delete(0, "end")
            self.address_entry.insert(0, self.settings_manager.shop_info.address)
            
            self.tax_number_entry.delete(0, "end")
            self.tax_number_entry.insert(0, self.settings_manager.shop_info.tax_number)

            # Display settings
            self.theme_var.set(self.settings_manager.display.theme)
            self.language_var.set(self.settings_manager.display.language)
            self.font_size_var.set(self.settings_manager.display.font_size)
            self.items_per_page_var.set(str(self.settings_manager.display.items_per_page))
            self.show_grid_var.set(self.settings_manager.display.show_grid)

            # Business settings
            self.currency_entry.delete(0, "end")
            self.currency_entry.insert(0, self.settings_manager.business.currency)
            
            self.currency_symbol_entry.delete(0, "end")
            self.currency_symbol_entry.insert(0, self.settings_manager.business.currency_symbol)
            
            self.tax_rate_entry.delete(0, "end")
            self.tax_rate_entry.insert(0, str(self.settings_manager.business.tax_rate))
            
            self.default_discount_entry.delete(0, "end")
            self.default_discount_entry.insert(0, str(self.settings_manager.business.default_discount))
            
            self.backup_interval_entry.delete(0, "end")
            self.backup_interval_entry.insert(0, str(self.settings_manager.business.backup_interval_days))
            
            self.low_stock_alert_var.set(self.settings_manager.business.low_stock_alert)

            self._update_font_size_label()
            
            logger.info("Settings loaded successfully")

        except Exception as e:
            logger.error(f"Error loading settings: {e}")
            messagebox.showerror("خطأ", f"حدث خطأ في تحميل الإعدادات: {e}")

    def _save_settings(self):
        """Save settings"""
        try:
            # Update shop info
            self.settings_manager.update_shop_info(
                name=self.shop_name_entry.get(),
                owner=self.owner_name_entry.get(),
                phone=self.phone_entry.get(),
                email=self.email_entry.get(),
                address=self.address_entry.get(),
                tax_number=self.tax_number_entry.get()
            )

            # Update display settings
            self.settings_manager.update_display_settings(
                theme=self.theme_var.get(),
                language=self.language_var.get(),
                font_size=self.font_size_var.get(),
                items_per_page=int(self.items_per_page_var.get()),
                show_grid=self.show_grid_var.get()
            )

            # Update business settings
            self.settings_manager.update_business_settings(
                currency=self.currency_entry.get(),
                currency_symbol=self.currency_symbol_entry.get(),
                tax_rate=float(self.tax_rate_entry.get() or 0),
                default_discount=float(self.default_discount_entry.get() or 0),
                backup_interval_days=int(self.backup_interval_entry.get() or 7),
                low_stock_alert=self.low_stock_alert_var.get()
            )

            messagebox.showinfo("نجح", "تم حفظ الإعدادات بنجاح!")
            logger.info("Settings saved successfully")

        except ValueError as e:
            messagebox.showerror("خطأ", "يرجى التأكد من صحة القيم المدخلة")
            logger.error(f"Invalid input values: {e}")
        except Exception as e:
            logger.error(f"Error saving settings: {e}")
            messagebox.showerror("خطأ", f"حدث خطأ في حفظ الإعدادات: {e}")

    def _on_theme_change(self, value):
        """Handle theme change"""
        try:
            self.theme_manager.switch_theme(value)
            messagebox.showinfo("تغيير السمة", "تم تغيير السمة بنجاح! يرجى إعادة تشغيل التطبيق لتطبيق التغييرات بالكامل.")
        except Exception as e:
            logger.error(f"Error changing theme: {e}")

    def _on_font_size_change(self, *args):
        """Handle font size change"""
        self._update_font_size_label()

    def _update_font_size_label(self):
        """Update font size label"""
        self.font_size_label.configure(text=str(self.font_size_var.get()))

    def _create_backup(self):
        """Create database backup"""
        try:
            from datetime import datetime
            import shutil
            
            backup_dir = Path("data/backups")
            backup_dir.mkdir(parents=True, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_name = f"backup_{timestamp}.db"
            backup_path = backup_dir / backup_name
            
            # Copy database file
            db_path = Path("data/database/shop.db")
            if db_path.exists():
                shutil.copy2(db_path, backup_path)
                messagebox.showinfo("نجح", f"تم إنشاء النسخة الاحتياطية بنجاح: {backup_name}")
                logger.info(f"Database backup created: {backup_path}")
            else:
                messagebox.showerror("خطأ", "لم يتم العثور على قاعدة البيانات")
                
        except Exception as e:
            logger.error(f"Error creating backup: {e}")
            messagebox.showerror("خطأ", f"حدث خطأ في إنشاء النسخة الاحتياطية: {e}")

    def _restore_backup(self):
        """Restore database backup"""
        try:
            backup_file = filedialog.askopenfilename(
                title="اختر ملف النسخة الاحتياطية",
                initialdir="data/backups",
                filetypes=[("Database files", "*.db"), ("All files", "*.*")]
            )
            
            if backup_file:
                import shutil
                
                # Confirm restore
                if messagebox.askyesno("تأكيد", "هل أنت متأكد من استعادة النسخة الاحتياطية؟ سيتم استبدال البيانات الحالية."):
                    db_path = Path("data/database/shop.db")
                    shutil.copy2(backup_file, db_path)
                    messagebox.showinfo("نجح", "تم استعادة النسخة الاحتياطية بنجاح! يرجى إعادة تشغيل التطبيق.")
                    logger.info(f"Database restored from: {backup_file}")
                    
        except Exception as e:
            logger.error(f"Error restoring backup: {e}")
            messagebox.showerror("خطأ", f"حدث خطأ في استعادة النسخة الاحتياطية: {e}")

    def _export_data(self):
        """Export data to Excel"""
        try:
            from datetime import datetime
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Export products
            ws_products = wb.active
            ws_products.title = "المنتجات"
            
            # Headers
            headers = ["ID", "الاسم", "الباركود", "الفئة", "السعر", "المخزون", "تاريخ الإضافة"]
            for col, header in enumerate(headers, 1):
                cell = ws_products.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Get products data
            products = self.db_manager.get_all_products()
            for row, product in enumerate(products, 2):
                ws_products.cell(row=row, column=1, value=product[0])  # ID
                ws_products.cell(row=row, column=2, value=product[1])  # Name
                ws_products.cell(row=row, column=3, value=product[2])  # Barcode
                ws_products.cell(row=row, column=4, value=product[3])  # Category
                ws_products.cell(row=row, column=5, value=product[4])  # Price
                ws_products.cell(row=row, column=6, value=product[5])  # Stock
                ws_products.cell(row=row, column=7, value=product[6])  # Date added
            
            # Export sales
            ws_sales = wb.create_sheet("المبيعات")
            sales_headers = ["ID", "التاريخ", "العميل", "المجموع", "الخصم", "الصافي"]
            for col, header in enumerate(sales_headers, 1):
                cell = ws_sales.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Get sales data
            sales = self.db_manager.get_all_sales()
            for row, sale in enumerate(sales, 2):
                ws_sales.cell(row=row, column=1, value=sale[0])  # ID
                ws_sales.cell(row=row, column=2, value=sale[1])  # Date
                ws_sales.cell(row=row, column=3, value=sale[2])  # Customer
                ws_sales.cell(row=row, column=4, value=sale[3])  # Total
                ws_sales.cell(row=row, column=5, value=sale[4])  # Discount
                ws_sales.cell(row=row, column=6, value=sale[5])  # Net
            
            # Save file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"data_export_{timestamp}.xlsx"
            export_path = Path("data/exports") / filename
            export_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(export_path)
            
            messagebox.showinfo("نجح", f"تم تصدير البيانات بنجاح: {filename}")
            logger.info(f"Data exported to: {export_path}")
            
        except Exception as e:
            logger.error(f"Error exporting data: {e}")
            messagebox.showerror("خطأ", f"حدث خطأ في تصدير البيانات: {e}")

    def _reset_settings(self):
        """Reset all settings to defaults"""
        try:
            if messagebox.askyesno("تأكيد", "هل أنت متأكد من إعادة تعيين جميع الإعدادات؟"):
                self.settings_manager.reset_to_defaults()
                self._load_settings()
                messagebox.showinfo("نجح", "تم إعادة تعيين الإعدادات بنجاح!")
                logger.info("Settings reset to defaults")
                
        except Exception as e:
            logger.error(f"Error resetting settings: {e}")
            messagebox.showerror("خطأ", f"حدث خطأ في إعادة تعيين الإعدادات: {e}")
